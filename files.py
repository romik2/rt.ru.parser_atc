import os
from openpyxl import load_workbook

def get_gata_workbook(pathFileExport):
    calls = []
    wb = load_workbook(pathFileExport)
    sheet = wb.active

    i = 2
    n = 4
    while True:
        calls.append([])
        for j in range(13):
            calls[i - 2].append(sheet.cell(row=i, column=j + 1).value)
        i += 1
        if (sheet.cell(row=i, column=1).value == None):
            break

    # os.remove(pathFileExport)
    return calls